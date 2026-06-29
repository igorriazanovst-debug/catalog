"""Предварительный разбор входящей сметы (xlsx) БЕЗ LLM.

Формат сметы НЕ фиксирован по колонкам. Разные заказчики присылают разные
шаблоны 44-ФЗ «Описание объекта закупки»:
  * шапка таблицы может стоять на любой строке (выше — титул/реквизиты);
  * колонки называются по-разному и бывают объединены (одна логическая
    колонка «Характеристики» раскрывается в несколько подколонок);
  * одна позиция занимает НЕСКОЛЬКО строк (первая строка — сама позиция,
    дальше идут её характеристики; № п/п и наименование заполнены только в
    первой строке, часто через вертикальное объединение ячеек);
  * код КТРУ/ОКПД2 бывает в отдельной колонке, а бывает зашит прямо в текст
    наименования («... Код: 32.99.53.130-00000136; Наименование: ...»);
  * снизу таблицы идут «подвалы» (Итого, требования к качеству/таре и т.п.).

Поэтому разбор — ЭВРИСТИЧЕСКИЙ, в два шага:
  1) найти строку-шапку и сопоставить колонки каноническим полям
     (`num`, `name`, `code`, `quantity`, `unit`, `price`, `total`,
      `char_name`, `char_value`, `char_unit`);
  2) сгруппировать строки в позиции (новая позиция = строка, где заполнен
     номер/наименование; последующие строки — её характеристики).

Результат — нормализованный список позиций + ДИАГНОСТИКА (какая строка
распознана шапкой, как легли колонки, какие предупреждения). Это «контракт»,
на который дальше опираются шаги сопоставления (позиция → 838 → товары →
цены) независимо от исходного шаблона. LLM здесь не используется — только
правила; подключение LLM — следующий этап.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl


# --- Канонические поля колонок ------------------------------------------------
# Правила сопоставления заголовка колонки полю. Порядок ВАЖЕН: более конкретные
# образцы идут раньше общих (например, «единица измерения ХАРАКТЕРИСТИКИ» должна
# пойматься раньше, чем просто «единица измерения» позиции). Каждое поле, кроме
# `char_value`, занимает не более одной колонки (первая совпавшая выигрывает).
_SINGLE_RULES: list[tuple[str, list[str]]] = [
    ("char_name", [r"наименование характеристик"]),
    ("char_unit", [r"единица измерения характеристик"]),
    ("name", [r"наименование товара", r"^наименование,? работ", r"^наименование$"]),
    # ВАЖНО: код должен начинаться со слова «код» — иначе колонка «Применение
    # характеристик из справочника КТРУ» ложно опознаётся как колонка кода.
    ("code", [r"^код\b", r"код позиции", r"код .{0,15}ктру", r"код .{0,15}окпд", r"^ктру/окпд", r"^окпд2?$"]),
    ("trademark", [r"товарный знак", r"товарного знака"]),
    ("quantity", [r"количеств", r"кол-?во"]),
    ("unit", [r"единиц[аы] измерени", r"ед\.? ?изм"]),  # после char_unit
    ("price", [r"цена за единиц", r"цена,? руб", r"^цена\b"]),
    ("total", [r"стоимость позиции", r"^стоимость", r"^сумма\b"]),
    ("num", [r"№ ?п/п", r"^№\b", r"^n п/п", r"^п/п$"]),
]

# Колонки-«значения характеристики». Их может быть несколько (текстовое
# значение + операторы >=/<= + Min/Max + «Конкретное значение»). При сборке
# характеристики значения этих колонок склеиваются в одну строку — так
# «>= 240» или «<= 1» собирается из колонок-операторов и Min/Max.
_VALUE_PATTERNS = [
    r"значение характеристик",
    r"текстовое описание значени",
    r"конкретное значение",
    r"^min$",
    r"^max$",
    r"^>",   # «>», «>, >=»
    r"^<",   # «<», «<, <=»
    r">=",
    r"<=",
]

# Маркеры «подвала» сметы — после первой позиции встреча любого из них завершает
# таблицу позиций (дальше идут требования к качеству/таре, итоги и т.п.).
_FOOTER_MARKERS = [
    "итого",
    "поставляемый товар должен",
    "требования к качеству",
    "требования к таре",
    "тип объекта закупки",
]

# Код КТРУ (напр. 32.99.53.130-00000136) и ОКПД2 (напр. 32.99.53.130).
_KTRU_RE = re.compile(r"\d{2}\.\d{2}\.\d{2}\.\d{3}-\d{8,}")
_OKPD2_RE = re.compile(r"\d{2}\.\d{2}\.\d{2}\.\d{3}(?!-)")
# Код, зашитый в текст наименования: «... Код: <код>; Наименование: ...».
_INLINE_CODE_RE = re.compile(r"код[:\s]+([0-9.\-]+)", re.IGNORECASE)


@dataclass
class Characteristic:
    name: str
    value: str = ""
    unit: str = ""


@dataclass
class EstimateItem:
    position: str | None = None          # № п/п как в смете (строкой)
    name: str = ""
    code_ktru: str | None = None
    code_okpd2: str | None = None
    code_raw: str | None = None          # исходный текст кода (как в смете)
    quantity: str | None = None
    unit: str | None = None
    price: str | None = None
    total: str | None = None
    characteristics: list[Characteristic] = field(default_factory=list)
    source_rows: list[int] = field(default_factory=list)  # 1-based номера строк
    warnings: list[str] = field(default_factory=list)


# --- Низкоуровневые помощники -------------------------------------------------

def _norm(v) -> str:
    """Нормализованный текст ячейки: схлопнутые пробелы/переводы строк, lower."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def _text(v) -> str:
    """Исходный текст ячейки без агрессивной нормализации (только strip)."""
    if v is None:
        return ""
    return str(v).strip()


def _grid(ws) -> list[list]:
    """Плотная сетка значений листа (natural: не верх-лево объединения = None)."""
    nrows, ncols = ws.max_row, ws.max_column
    grid = [[None] * ncols for _ in range(nrows)]
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            grid[r - 1][c - 1] = ws.cell(row=r, column=c).value
    return grid


def _hfill(grid: list[list], ws) -> list[list]:
    """Сетка с ГОРИЗОНТАЛЬНЫМ растеканием объединённых ячеек (только в пределах
    верхней строки объединения). Нужна для чтения шапки: объединённый заголовок
    «Характеристики» на 5 колонок становится виден в каждой из них. Вертикально
    НЕ растекаем — иначе номер/наименование позиции «протекли» бы в строки её
    характеристик и сломали группировку."""
    g = [row[:] for row in grid]
    for mr in ws.merged_cells.ranges:
        tl = grid[mr.min_row - 1][mr.min_col - 1]
        for c in range(mr.min_col, mr.max_col + 1):
            g[mr.min_row - 1][c - 1] = tl
    return g


def _matches(text: str, patterns: list[str]) -> bool:
    return any(re.search(p, text) for p in patterns)


def _classify_columns(header_texts: list[str]) -> tuple[dict, list[int]]:
    """header_texts[i] — нормализованный заголовок колонки i (0-based).
    Возвращает (single_map: field->col_index, value_cols: list[col_index])."""
    single: dict[str, int] = {}
    value_cols: list[int] = []
    for col, htext in enumerate(header_texts):
        if not htext:
            continue
        if _matches(htext, _VALUE_PATTERNS):
            value_cols.append(col)
            # «Конкретное значение»/«Min»/«Max» — это только значение, дальше не
            # классифицируем как одиночное поле.
            continue
        for fieldname, patterns in _SINGLE_RULES:
            if fieldname in single:
                continue
            if _matches(htext, patterns):
                single[fieldname] = col
                break
    return single, value_cols


def _header_score(header_texts: list[str]) -> int:
    """Сколько разных канонических полей распознаётся в строке-кандидате шапки."""
    single, value_cols = _classify_columns(header_texts)
    return len(single) + (1 if value_cols else 0)


def _is_legend_row(row: list) -> bool:
    """Строка-«легенда» нумерации колонок: 1 2 3 ... — её надо пропустить."""
    nums = [_text(x) for x in row if _text(x)]
    if len(nums) < 3:
        return False
    return all(re.fullmatch(r"\d+", n) for n in nums)


def _looks_like_item_start(row: list, single: dict) -> bool:
    """Строка открывает новую позицию: заполнен номер (целое) или, если колонки
    номера нет, заполнено наименование товара."""
    num_col = single.get("num")
    if num_col is not None:
        if re.fullmatch(r"\d+", _text(row[num_col])):
            return True
        # Номер может быть пустым у валидной позиции — тогда смотрим имя.
    name_col = single.get("name")
    if name_col is not None and _text(row[name_col]):
        # Но не путаем с характеристикой: у строки-характеристики имя товара
        # пустое (заполнено char_name). Если num-колонка есть и пуста — это,
        # скорее всего, характеристика, не позиция.
        if num_col is None:
            return True
    return False


def _extract_codes(raw: str) -> tuple[str | None, str | None]:
    """Из текста кода вытащить (ктру, окпд2)."""
    if not raw:
        return None, None
    ktru = _KTRU_RE.search(raw)
    okpd = _OKPD2_RE.search(raw)
    ktru_code = ktru.group(0) if ktru else None
    okpd_code = okpd.group(0) if okpd else None
    # ОКПД2 — это префикс кода КТРУ до дефиса. Если ОКПД2 отдельно не нашли, но
    # есть КТРУ — выводим ОКПД2 из него (полезно для сопоставления по коду с 838).
    if ktru_code and not okpd_code:
        okpd_code = ktru_code.split("-")[0]
    return ktru_code, okpd_code


def _split_name_and_code(name_cell: str) -> tuple[str, str | None]:
    """Если код зашит в наименование («Имя\nКод: ...; Наименование: ...; Версия:»),
    вернуть (чистое имя, сырой код). Иначе (имя как есть, None)."""
    if not name_cell:
        return "", None
    # Берём первую содержательную строку как наименование.
    first_line = name_cell.splitlines()[0].strip()
    m = _INLINE_CODE_RE.search(name_cell)
    code_raw = m.group(1).rstrip(";.") if m else None
    name = first_line if first_line else name_cell.strip()
    return name, code_raw


# --- Основной разбор ----------------------------------------------------------

def _find_header(hgrid: list[list]) -> tuple[int, list[str], int]:
    """Найти строку-шапку (по максимальному числу распознанных полей) и собрать
    итоговые заголовки колонок. Если соседняя снизу строка тоже «шапочная»
    (подзаголовки объединённого блока), её текст приклеивается к заголовкам.

    Возвращает (header_row, headers, header_end): индекс строки-шапки (0-based),
    собранные заголовки по колонкам и индекс ПОСЛЕДНЕЙ строки шапки (0-based) —
    данные начинаются ниже header_end. Если шапка не найдена — (-1, [], -1)."""
    best_row, best_score = -1, 0
    per_row_texts = [[_norm(x) for x in row] for row in hgrid]
    for r, texts in enumerate(per_row_texts):
        score = _header_score(texts)
        if score > best_score:
            best_score, best_row = score, r
    if best_row < 0:
        return -1, [], -1

    headers = list(per_row_texts[best_row])
    # Подзаголовки: следующая строка тоже похожа на шапку (>=2 полей) — склеиваем
    # поколоночно. Это раскрывает «Характеристики» → «наименование/значение/ед.».
    nxt = best_row + 1
    if nxt < len(per_row_texts) and _header_score(per_row_texts[nxt]) >= 2:
        for c, sub in enumerate(per_row_texts[nxt]):
            if sub and sub != headers[c]:
                headers[c] = f"{headers[c]} {sub}".strip()
        header_end = nxt
    else:
        header_end = best_row
    return best_row, headers, header_end


def parse_worksheet(ws) -> dict:
    """Разобрать один лист. Возвращает словарь с позициями и диагностикой."""
    grid = _grid(ws)
    hgrid = _hfill(grid, ws)

    header_row, headers, header_end = _find_header(hgrid)
    warnings: list[str] = []
    if header_row < 0:
        return {
            "sheet": ws.title,
            "header_row": None,
            "columns": {},
            "value_columns": [],
            "column_headers": {},
            "items": [],
            "warnings": ["Не удалось распознать строку-шапку таблицы."],
        }

    single, value_cols = _classify_columns(headers)

    # Колоночные предупреждения о пропусках, важных для дальнейших шагов.
    if "name" not in single:
        warnings.append("Колонка с наименованием товара не распознана.")
    if "code" not in single:
        warnings.append(
            "Отдельной колонки кода КТРУ/ОКПД2 нет — коды будем искать в тексте наименования."
        )
    if "quantity" not in single:
        warnings.append("Колонка «Количество» не распознана.")

    items: list[EstimateItem] = []
    cur: EstimateItem | None = None
    started = False

    for r in range(header_end + 1, len(grid)):
        row = grid[r]
        rownum = r + 1  # 1-based для диагностики
        if all(_text(x) == "" for x in row):
            continue
        joined = " ".join(_norm(x) for x in row)
        if started and _matches(joined, _FOOTER_MARKERS):
            break
        if _is_legend_row(row):
            continue

        if _looks_like_item_start(row, single):
            cur = _new_item(row, rownum, single, value_cols)
            items.append(cur)
            started = True
            _add_characteristic(cur, row, single, value_cols)
        elif cur is not None:
            cur.source_rows.append(rownum)
            _add_characteristic(cur, row, single, value_cols)
        # строки до первой позиции игнорируем

    # Постобработка позиций: дотащить код из наименования, проставить warnings.
    for it in items:
        if not it.code_ktru and not it.code_okpd2 and it.code_raw:
            it.code_ktru, it.code_okpd2 = _extract_codes(it.code_raw)
        if not it.code_ktru and not it.code_okpd2:
            it.warnings.append("Код КТРУ/ОКПД2 не распознан.")
        if not it.name:
            it.warnings.append("Наименование пустое.")

    return {
        "sheet": ws.title,
        "header_row": header_row + 1,
        "columns": {f: _col_letter(c) for f, c in single.items()},
        "value_columns": [_col_letter(c) for c in value_cols],
        "column_headers": {
            _col_letter(c): headers[c] for c in range(len(headers)) if headers[c]
        },
        "items": [asdict(it) for it in items],
        "warnings": warnings,
    }


def _new_item(row, rownum, single, value_cols) -> EstimateItem:
    it = EstimateItem(source_rows=[rownum])

    num_col = single.get("num")
    if num_col is not None:
        it.position = _text(row[num_col]) or None

    name_col = single.get("name")
    raw_name = _text(row[name_col]) if name_col is not None else ""
    name, inline_code = _split_name_and_code(raw_name)
    it.name = name

    code_col = single.get("code")
    code_raw = _text(row[code_col]) if code_col is not None else ""
    ktru, okpd = _extract_codes(code_raw) if code_raw else (None, None)
    # Колонка кода может отсутствовать или содержать мусор («Используется
    # справочник КТРУ ЕИС») — тогда берём код, зашитый в текст наименования.
    if not ktru and not okpd and inline_code:
        code_raw = inline_code
        ktru, okpd = _extract_codes(code_raw)
    it.code_raw = code_raw or None
    it.code_ktru, it.code_okpd2 = ktru, okpd

    for fieldname in ("quantity", "unit", "price", "total"):
        col = single.get(fieldname)
        if col is not None:
            setattr(it, fieldname, _text(row[col]) or None)

    return it


def _add_characteristic(it: EstimateItem, row, single, value_cols):
    """Добавить характеристику, если в строке заполнено имя характеристики."""
    char_col = single.get("char_name")
    if char_col is None:
        return
    cname = _text(row[char_col])
    if not cname:
        return
    parts = [_text(row[c]) for c in value_cols if _text(row[c])]
    value = " ".join(parts)
    unit_col = single.get("char_unit")
    cunit = _text(row[unit_col]) if unit_col is not None else ""
    it.characteristics.append(Characteristic(name=cname, value=value, unit=cunit))


def _col_letter(idx0: int) -> str:
    """0-based индекс колонки -> буква (0->A, 26->AA)."""
    s = ""
    n = idx0 + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_estimate(path: str | Path) -> dict:
    """Разобрать книгу: выбрать лист с лучшей шапкой и вернуть его разбор.
    В диагностику кладём список всех листов и какой выбран."""
    wb = openpyxl.load_workbook(path, data_only=True)
    best = None
    best_score = -1
    for ws in wb.worksheets:
        hgrid = _hfill(_grid(ws), ws)
        hr, _h, _he = _find_header(hgrid)
        score = _header_score([_norm(x) for x in hgrid[hr]]) if hr >= 0 else -1
        if score > best_score:
            best_score = score
            best = ws
    if best is None:
        return {"file": str(path), "sheets": wb.sheetnames, "result": None,
                "warnings": ["В книге нет листов."]}
    result = parse_worksheet(best)
    result["file"] = str(path)
    result["sheets"] = wb.sheetnames
    return result
