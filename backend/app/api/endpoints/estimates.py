"""API входящих смет: загрузка (фон) → подбор → просмотр/правка → экспорт.

Пайплайн загрузки (как у импорта товаров — длинная операция с LLM, поэтому в фоне):
  POST /api/estimates/upload  — принимает xlsx, сразу отдаёт {job_id}; разбор+
    подбор+запись идут фоном, прогресс/итог — через GET /api/jobs/{id}.
  GET  /api/estimates              — список смет.
  GET  /api/estimates/{id}         — смета с позициями (товар/поставщик/цена).
  GET  /api/estimates/{id}/items/{item_id}/candidates — варианты товаров для строки.
  POST /api/estimates/{id}/items/{item_id}/choose      — ручной выбор товара/поставщика.
  GET  /api/estimates/{id}/export  — выгрузка сметы в xlsx.
  DELETE /api/estimates/{id}       — удалить смету.
"""

import asyncio
import base64
import io

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db, async_session
from app.services.estimate_parser import parse_estimate
from app.services.estimate_service import EstimateMatcher
from app.services.mapping_service import get_embedding_model
from app.services.jobs import jobs, run_job

router = APIRouter(prefix="/api/estimates", tags=["estimates"])

_XLSX_EXT = (".xlsx", ".xlsm", ".xltx", ".xltm")


@router.post("/upload")
async def upload_estimate(
    file: UploadFile = File(...),
    name: str = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Загрузить смету (xlsx): РАЗОБРАТЬ и сохранить распознанные строки (без
    подбора). Подбор запускается отдельно — авто (classify) или построчно. Сам
    файл сохраняется (для аннотированного экспорта)."""
    if not file.filename or not file.filename.lower().endswith(_XLSX_EXT):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате xlsx")
    content = await file.read()
    try:
        parsed = parse_estimate(io.BytesIO(content), display_name=file.filename)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось разобрать xlsx: {e}")
    if not parsed.get("items"):
        raise HTTPException(
            status_code=400,
            detail="В файле не найдено позиций сметы: " + "; ".join(parsed.get("warnings", [])))

    est_name = (name or "").strip() or file.filename
    b64 = base64.b64encode(content).decode("ascii")
    matcher = EstimateMatcher(db)
    saved = await matcher.create_estimate_from_parsed(
        est_name, parsed, file.filename, b64)
    return {"estimate_id": saved["estimate_id"], "name": est_name,
            "positions": saved["items"], "warnings": parsed.get("warnings", [])}


@router.post("/{estimate_id}/classify")
async def classify_estimate(
    estimate_id: int,
    use_llm: bool = Form(True),
    provider: str = Form(None),
    decompose: bool = Form(True),
    price_basis: str = Form("cost"),
    db: AsyncSession = Depends(get_db),
):
    """Авто-классификация всей сметы (фон): подбор по всем строкам + перезапись."""
    if price_basis not in ("cost", "retail"):
        raise HTTPException(status_code=400, detail="price_basis: cost|retail")
    await _estimate_or_404(db, estimate_id)
    prov = None if (provider or "").strip() in ("", "default") else provider.strip()

    cnt = await db.execute(
        text("SELECT count(*) FROM estimate_items WHERE estimate_id = :id"),
        {"id": estimate_id})
    job = jobs.create("estimate")
    job.total = int(cnt.scalar() or 0)

    async def body(job):
        await asyncio.to_thread(get_embedding_model)
        async with async_session() as bg_db:
            matcher = EstimateMatcher(bg_db, price_basis=price_basis)
            return await matcher.classify_estimate(
                estimate_id, use_llm=use_llm, provider=prov, decompose=decompose,
                progress=lambda p, t: job.set_progress(p, t))

    run_job(job, body)
    return {"job_id": job.id, "estimate_id": estimate_id}


@router.post("/{estimate_id}/items/{item_id}/classify")
async def classify_one_item(
    estimate_id: int, item_id: int,
    use_llm: bool = False,
    provider: str = None,
    db: AsyncSession = Depends(get_db),
):
    """Классифицировать ОДНУ строку (ручной режим): без LLM (use_llm=false) или с
    LLM. Подбирает стандарт 838 и лучший товар, обновляет позицию."""
    prov = None if (provider or "").strip() in ("", "default") else provider.strip()
    # Прогрев модели эмбеддингов в потоке (первый вызов после рестарта ~минуту),
    # чтобы не блокировать event loop остальных запросов.
    await asyncio.to_thread(get_embedding_model)
    matcher = EstimateMatcher(db)
    try:
        res = await matcher.classify_item(estimate_id, item_id,
                                          use_llm=use_llm, provider=prov)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    std = res.get("standard")
    offer = res.get("chosen_offer")
    return {
        "match_method": res.get("match_method"),
        "match_reason": res.get("match_reason"),
        "standard_id": std["standard_id"] if std else None,
        "standard_name": std.get("standard_name") if std else None,
        "product_id": offer["product_id"] if offer else None,
        "unit_price": res.get("unit_price"),
        "total_price": res.get("total_price"),
    }


@router.get("")
async def list_estimates(db: AsyncSession = Depends(get_db)):
    res = await db.execute(text("""
        SELECT e.id, e.name, e.description, e.total_amount, e.created_at,
               COUNT(i.id) AS items,
               COUNT(i.id) FILTER (WHERE i.product_id IS NOT NULL) AS matched
        FROM estimates e
        LEFT JOIN estimate_items i ON i.estimate_id = e.id
        GROUP BY e.id
        ORDER BY e.created_at DESC, e.id DESC
    """))
    items = [{
        "id": r[0], "name": r[1], "description": r[2],
        "total_amount": float(r[3]) if r[3] is not None else 0.0,
        "created_at": r[4].isoformat() if r[4] else None,
        "items": r[5] or 0, "matched": r[6] or 0,
    } for r in res.fetchall()]
    return {"items": items}


async def _estimate_or_404(db, estimate_id: int):
    res = await db.execute(
        text("SELECT id, name, description, total_amount, created_at "
             "FROM estimates WHERE id = :id"), {"id": estimate_id})
    row = res.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="смета не найдена")
    return row


@router.get("/{estimate_id}")
async def get_estimate(estimate_id: int, db: AsyncSession = Depends(get_db)):
    e = await _estimate_or_404(db, estimate_id)
    res = await db.execute(text("""
        SELECT i.id, i.source_name, i.source_description, i.group_name, i.unit,
               i.match_method, i.match_reason, i.quantity, i.unit_price, i.total_price,
               i.standard_id, st.item_name, st.full_code,
               i.product_id, p.name, p.sku, p.description,
               i.supplier_id, s.name
        FROM estimate_items i
        LEFT JOIN industry_standards st ON st.id = i.standard_id
        LEFT JOIN products p ON p.id = i.product_id
        LEFT JOIN suppliers s ON s.id = i.supplier_id
        WHERE i.estimate_id = :id
        ORDER BY i.id
    """), {"id": estimate_id})
    items = [{
        "id": r[0], "source_name": r[1], "source_description": r[2],
        "group_name": r[3], "unit": r[4],
        "match_method": r[5], "match_reason": r[6],
        "quantity": float(r[7]) if r[7] is not None else None,
        "unit_price": float(r[8]) if r[8] is not None else None,
        "total_price": float(r[9]) if r[9] is not None else None,
        "standard_id": r[10], "standard_name": r[11], "full_code": r[12],
        "product_id": r[13], "product_name": r[14], "sku": r[15],
        "product_description": r[16],
        "supplier_id": r[17], "supplier_name": r[18],
    } for r in res.fetchall()]
    return {
        "id": e[0], "name": e[1], "description": e[2],
        "total_amount": float(e[3]) if e[3] is not None else 0.0,
        "created_at": e[4].isoformat() if e[4] else None,
        "items": items,
    }


@router.get("/{estimate_id}/items/{item_id}/candidates")
async def item_candidates(estimate_id: int, item_id: int,
                          db: AsyncSession = Depends(get_db)):
    """Варианты товаров для строки (по её стандарту 838) — для ручного выбора."""
    res = await db.execute(
        text("SELECT standard_id FROM estimate_items "
             "WHERE id = :iid AND estimate_id = :eid"),
        {"iid": item_id, "eid": estimate_id})
    row = res.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="позиция не найдена")
    if row[0] is None:
        return {"candidates": []}
    matcher = EstimateMatcher(db)
    offers = await matcher.offers_for_standard(row[0])
    return {"candidates": offers}


@router.post("/{estimate_id}/items/{item_id}/choose")
async def choose_item(estimate_id: int, item_id: int,
                      product_id: int, supplier_id: int,
                      price_basis: str = "cost",
                      db: AsyncSession = Depends(get_db)):
    """Ручной выбор товара/поставщика для строки: пересчёт цены и итога сметы."""
    if price_basis not in ("cost", "retail"):
        raise HTTPException(status_code=400, detail="price_basis: cost|retail")
    res = await db.execute(
        text("SELECT quantity FROM estimate_items "
             "WHERE id = :iid AND estimate_id = :eid"),
        {"iid": item_id, "eid": estimate_id})
    row = res.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="позиция не найдена")
    qty = float(row[0]) if row[0] is not None else 1.0

    price_col = "cost_price" if price_basis == "cost" else "retail_price"
    pres = await db.execute(
        text(f"SELECT {price_col} FROM supplier_products "
             "WHERE product_id = :pid AND supplier_id = :supid"),
        {"pid": product_id, "supid": supplier_id})
    prow = pres.fetchone()
    if not prow:
        raise HTTPException(status_code=400, detail="нет такого предложения поставщика")
    unit_price = float(prow[0]) if prow[0] is not None else 0.0
    total_price = unit_price * qty

    await db.execute(
        text("""UPDATE estimate_items
                SET product_id = :pid, supplier_id = :supid,
                    unit_price = :uprice, total_price = :tprice,
                    match_method = 'manual',
                    match_reason = 'ручной выбор товара/поставщика'
                WHERE id = :iid"""),
        {"pid": product_id, "supid": supplier_id, "uprice": unit_price,
         "tprice": total_price, "iid": item_id})
    # Пересчёт итога сметы.
    await db.execute(
        text("""UPDATE estimates SET total_amount =
                 (SELECT COALESCE(SUM(total_price), 0) FROM estimate_items
                  WHERE estimate_id = :eid)
                WHERE id = :eid"""),
        {"eid": estimate_id})
    await db.commit()
    return {"status": "ok", "unit_price": unit_price, "total_price": total_price}


@router.delete("/{estimate_id}")
async def delete_estimate(estimate_id: int, db: AsyncSession = Depends(get_db)):
    await _estimate_or_404(db, estimate_id)
    await db.execute(text("DELETE FROM estimates WHERE id = :id"), {"id": estimate_id})
    await db.commit()
    return {"status": "deleted"}


OUR_COLUMNS = ["Артикул", "Наименование (подбор)", "Описание (подбор)",
               "Цена за ед. (себест.)", "Итого"]


@router.get("/{estimate_id}/export")
async def export_estimate(estimate_id: int, db: AsyncSession = Depends(get_db)):
    """Экспорт: в ОРИГИНАЛЬНЫЙ файл сметы дописываем наши колонки (Артикул,
    Наименование, Описание, цена за ед, итого) и заполняем подбором. Плюс
    отдельный лист «Подбор» с полной плоской детализацией и итогами с НДС.
    Если исходный файл не сохранён — отдаём только лист «Подбор»."""
    import openpyxl
    from openpyxl.styles import Font

    e = await db.execute(
        text("SELECT name, source_file_b64, sheet_name, header_row "
             "FROM estimates WHERE id = :id"), {"id": estimate_id})
    erow = e.fetchone()
    if not erow:
        raise HTTPException(status_code=404, detail="смета не найдена")
    detail = await get_estimate(estimate_id, db)
    vat_res = await db.execute(
        text("SELECT value FROM system_settings WHERE key = 'vat_rate'"))
    try:
        vat = float(vat_res.scalar() or 0.0)
    except (TypeError, ValueError):
        vat = 0.0
    items = detail["items"]

    # --- Аннотируем исходный файл (если сохранён) ---
    wb = None
    if erow[1]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(erow[1])))
            ws = wb[erow[2]] if erow[2] and erow[2] in wb.sheetnames else wb.active
            header_row = erow[3] or 1
            start = ws.max_column + 2
            for j, h in enumerate(OUR_COLUMNS):
                cell = ws.cell(row=header_row, column=start + j, value=h)
                cell.font = Font(bold=True)
            # Группируем позиции по строке исходного файла (набор → несколько).
            # source_row нет в detail — дочитаем привязку id→строка одним запросом.
            by_row: dict[int, list] = {}
            rows_res = await db.execute(
                text("SELECT id, source_row FROM estimate_items WHERE estimate_id = :id"),
                {"id": estimate_id})
            srow = {r[0]: r[1] for r in rows_res.fetchall()}
            for it in items:
                sr = srow.get(it["id"])
                if sr:
                    by_row.setdefault(sr, []).append(it)
            for sr, its in by_row.items():
                if len(its) == 1:
                    it = its[0]
                    vals = [it["sku"] or "", it["product_name"] or "",
                            it["product_description"] or "", it["unit_price"],
                            it["total_price"]]
                else:  # набор: агрегат
                    names = "; ".join(x["product_name"] or x["source_name"] or "" for x in its)
                    total = sum(x["total_price"] or 0 for x in its)
                    vals = ["", f"Набор: {len(its)} поз.", names, "", round(total, 2)]
                for j, v in enumerate(vals):
                    ws.cell(row=sr, column=start + j, value=v)
        except Exception:  # noqa: BLE001
            wb = None  # если оригинал не открылся — упадём на лист «Подбор»

    if wb is None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # --- Лист «Подбор»: полная детализация ---
    ws2 = wb.create_sheet("Подбор")
    head = ["№", "Наименование (смета)", "Описание (смета)", "Набор",
            "Позиция 838", "Код 838", "Артикул", "Наименование (подбор)",
            "Описание (подбор)", "Поставщик", "Кол-во", "Ед.",
            "Цена за ед. (себест.)", "Итого"]
    ws2.append(head)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for n, it in enumerate(items, 1):
        ws2.append([
            n, it["source_name"], it["source_description"] or "",
            it["group_name"] or "", it["standard_name"] or "", it["full_code"] or "",
            it["sku"] or "", it["product_name"] or "", it["product_description"] or "",
            it["supplier_name"] or "", it["quantity"], it["unit"] or "",
            it["unit_price"], it["total_price"],
        ])
    subtotal = detail["total_amount"]
    vat_amount = round(subtotal * vat, 2)
    ws2.append([])
    ws2.append(["", "", "", "", "", "", "", "", "", "", "", "", "Итого:", subtotal])
    ws2.append(["", "", "", "", "", "", "", "", "", "", "", "",
                f"НДС {round(vat*100)}%:", vat_amount])
    ws2.append(["", "", "", "", "", "", "", "", "", "", "", "", "Всего с НДС:",
                round(subtotal + vat_amount, 2)])
    for off in (0, 1, 2):
        ws2[ws2.max_row - off][12].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"estimate_{estimate_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
