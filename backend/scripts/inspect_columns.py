"""Инспектор структуры файла-источника (xlsx/csv) для заполнения кодов КТРУ/ОКПД2.

Цель: понять, в каких колонках 838.xlsx (справочник) и/или прайса поставщика
лежат коды КТРУ/ОКПД2, чтобы написать корректный скрипт их проставления в
`industry_standards` / `products`. Скрипт НИЧЕГО не меняет — только печатает:
  * листы/размерность (для xlsx);
  * по каждой колонке — предполагаемый заголовок, примеры значений и СКОЛЬКО
    ячеек похожи на код КТРУ (NN.NN.NN.NNN-NNNNNNNN) и ОКПД2 (NN.NN.NN.NNN);
  * сводку: какие колонки выглядят как колонки кода.

Запуск:
    python scripts/inspect_columns.py ../data/input/838.xlsx
    python scripts/inspect_columns.py ../data/input/838.xlsx --sheet Лист1 --rows 8
    python scripts/inspect_columns.py ../data/input/price.csv
"""

import argparse
import sys
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from app.services.estimate_parser import _KTRU_RE, _OKPD2_RE  # noqa: E402


def _classify_codes(values: list[str]) -> tuple[int, int, list[str]]:
    """(сколько КТРУ, сколько ОКПД2, примеры совпадений) по списку значений."""
    ktru = okpd = 0
    samples = []
    for v in values:
        if not v:
            continue
        if _KTRU_RE.search(v):
            ktru += 1
            if len(samples) < 3:
                samples.append(_KTRU_RE.search(v).group(0))
        elif _OKPD2_RE.search(v):
            okpd += 1
            if len(samples) < 3:
                samples.append(_OKPD2_RE.search(v).group(0))
    return ktru, okpd, samples


def _report_columns(headers: list[str], columns: list[list[str]], rows_preview: int):
    code_cols = []
    for idx, (hdr, col_vals) in enumerate(zip(headers, columns)):
        nonempty = [v for v in col_vals if v]
        ktru, okpd, samples = _classify_codes(col_vals)
        sample_vals = " | ".join(nonempty[:rows_preview])[:120]
        flag = ""
        if ktru or okpd:
            flag = f"  <== КТРУ:{ktru} ОКПД2:{okpd} напр. {samples}"
            code_cols.append((idx, hdr, ktru, okpd))
        print(f"  [{idx}] {hdr[:45]!r:48} непустых={len(nonempty):<5} {sample_vals}{flag}")
    print()
    if code_cols:
        print("ПОХОЖЕ НА КОЛОНКИ КОДА:")
        for idx, hdr, ktru, okpd in code_cols:
            print(f"   колонка [{idx}] {hdr!r}: КТРУ={ktru}, ОКПД2={okpd}")
    else:
        print("Колонок с кодами КТРУ/ОКПД2 НЕ обнаружено.")


def inspect_xlsx(path: Path, sheet: str | None, rows_preview: int):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f"Листы: {wb.sheetnames}")
    targets = [sheet] if sheet else wb.sheetnames
    for sname in targets:
        ws = wb[sname]
        print("=" * 100)
        print(f"ЛИСТ: {sname}  размер: {ws.max_row}x{ws.max_column}")
        # Заголовки: первая строка с >=2 непустыми текстовыми ячейками.
        all_rows = [[("" if c is None else str(c).strip()) for c in row]
                    for row in ws.iter_rows(values_only=True)]
        if not all_rows:
            print("  (пусто)")
            continue
        ncols = max(len(r) for r in all_rows)
        all_rows = [r + [""] * (ncols - len(r)) for r in all_rows]
        header_idx = 0
        for i, r in enumerate(all_rows[:10]):
            if sum(1 for c in r if c) >= 2:
                header_idx = i
                break
        headers = [h or f"col{j}" for j, h in enumerate(all_rows[header_idx])]
        columns = [[all_rows[ri][cj] for ri in range(header_idx + 1, len(all_rows))]
                   for cj in range(ncols)]
        print(f"  (строка-заголовок: {header_idx + 1})")
        _report_columns(headers, columns, rows_preview)


def inspect_csv(path: Path, rows_preview: int):
    import pandas as pd
    df = None
    for sep in (None, ",", ";", "\t"):
        try:
            df = pd.read_csv(path, dtype=str, sep=sep, engine="python",
                             keep_default_na=False)
            if df.shape[1] > 1:
                break
        except Exception:
            continue
    if df is None:
        print("Не удалось прочитать CSV.", file=sys.stderr)
        return
    print("=" * 100)
    print(f"CSV: {df.shape[0]} строк, {df.shape[1]} колонок")
    headers = list(df.columns)
    columns = [df[h].fillna("").astype(str).tolist() for h in headers]
    _report_columns(headers, columns, rows_preview)


def main():
    ap = argparse.ArgumentParser(description="Инспектор колонок (поиск кодов КТРУ/ОКПД2)")
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None, help="конкретный лист (xlsx)")
    ap.add_argument("--rows", type=int, default=4, help="сколько примеров значений на колонку")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"Файл не найден: {path}", file=sys.stderr)
        raise SystemExit(1)

    print(f"ФАЙЛ: {path}")
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        inspect_xlsx(path, args.sheet, args.rows)
    elif path.suffix.lower() in {".csv", ".txt"}:
        inspect_csv(path, args.rows)
    else:
        print(f"Неизвестный формат: {path.suffix}", file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
