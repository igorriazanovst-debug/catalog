"""Парсер Приказа 838 (xlsx) -> плоский JSON позиций.

Иерархия берётся из КОДА в колонке A (надёжно), а не из текстовых заголовков:
  - 2-уровневый код (напр. "2.17") = позиция уровня раздела «по предметной
    области» (общая для всех кабинетов раздела);
  - 3-уровневый код (напр. "2.14.137") = раздел 2 / подраздел 14 / пункт 137.

Имена раздела/подраздела/части берутся из заголовков «Раздел/Подраздел/Часть».
Захватываются ВСЕ строки-пункты независимо от наличия над ними заголовка
«Основное/Дополнительное оборудование» (из-за этой зависимости старый парсер
терял ~489 позиций).

Выход: data/output/order_838_tree.json со структурой:
  {"metadata": {"order": "838"}, "positions": [ {position}, ... ]}
position: full_code, section_code, section_name, subsection_code,
          subsection_name, part_name, equipment_type, item_name, scope

Запуск:
    python scripts/parse_order_838.py            # из data/input/838.xlsx
    python scripts/parse_order_838.py --xlsx /path/838.xlsx --out /path/out.json
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_XLSX = REPO_ROOT / "data" / "input" / "838.xlsx"
DEFAULT_OUT = REPO_ROOT / "data" / "output" / "order_838_tree.json"

SECTION_RE = re.compile(r"^Раздел\s+(\d+)\.\s*(.*)$")
SUBSECTION_RE = re.compile(r"^Подраздел\s+(\d+)\.\s*(.*)$")
PART_RE = re.compile(r"^Часть\s+(\d+)\.\s*(.*)$")
CODE_RE = re.compile(r"^(\d+(?:\.\d+)+)\.?$")
EQUIP_TYPES = ("Основное оборудование", "Дополнительное вариативное оборудование")

GENERIC_SUBSECTION = "По предметной области (общее для раздела)"


def parse(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    section_names = {}          # "2" -> "Комплекс оснащения предметных кабинетов..."
    subsection_names = {}       # ("2","14") -> "Кабинет физики"
    cur_section = None          # code
    cur_subsection = None       # code
    cur_part = None             # name
    cur_etype = None            # "Основное оборудование" / ...

    positions = []

    for a, b in ws.iter_rows(values_only=True):
        col_a = ("" if a is None else str(a)).strip()
        col_b = ("" if b is None else str(b)).strip()
        if not col_a:
            continue

        m = SECTION_RE.match(col_a)
        if m:
            cur_section = m.group(1)
            section_names[cur_section] = m.group(2).strip()
            cur_subsection = None
            cur_part = None
            cur_etype = None
            continue

        m = SUBSECTION_RE.match(col_a)
        if m and cur_section:
            cur_subsection = m.group(1)
            name = m.group(2).strip()
            if name:
                subsection_names[(cur_section, cur_subsection)] = name
            cur_part = None
            cur_etype = None
            continue

        m = PART_RE.match(col_a)
        if m:
            cur_part = m.group(2).strip()
            cur_etype = None
            continue

        if any(col_a.startswith(t) for t in EQUIP_TYPES):
            cur_etype = "Основное" if col_a.startswith(EQUIP_TYPES[0]) else "Дополнительное"
            continue

        m = CODE_RE.match(col_a)
        if m and col_b:
            code = m.group(1)
            parts = code.split(".")
            sec = parts[0]
            if len(parts) == 2:
                sub = None
                sub_name = GENERIC_SUBSECTION
            else:
                sub = parts[1]
                sub_name = subsection_names.get((sec, sub)) or (
                    subsection_names.get((cur_section, cur_subsection))
                    if cur_subsection else None
                )
            positions.append({
                "full_code": code,
                "section_code": sec,
                "section_name": section_names.get(sec, ""),
                "subsection_code": sub,
                "subsection_name": sub_name or "",
                "part_name": cur_part,
                "equipment_type": cur_etype,
                "item_name": col_b,
                "scope": "section_generic" if sub is None else "cabinet",
            })

    return {"metadata": {"order": "838"}, "positions": positions}


def main():
    ap = argparse.ArgumentParser(description="Парсер Приказа 838 (xlsx) -> JSON")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"Файл не найден: {xlsx}")
        raise SystemExit(1)

    tree = parse(xlsx)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(tree, f, ensure_ascii=False, indent=2)
    print(f"[OK] Позиций: {len(tree['positions'])} -> {out}")


if __name__ == "__main__":
    main()
